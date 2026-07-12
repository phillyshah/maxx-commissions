#!/usr/bin/env python3
"""Structural render tests for the commission PDF pipeline.

These assert on the intermediate .xlsx that generate_pdfs hands to LibreOffice,
so they run anywhere (no soffice/rendering required). They are self-contained:
all fixtures are synthesized, nothing external is needed.

Run:  python tests/test_pdf_render.py     (exit 0 = pass)

Covers the v2.2 fixes:
  * logo is always inserted at a fixed 220x115 @ B1, never the (Excel-inflated)
    source size, so it can't overlap/erase the left column headers (row 5)
  * all B5:J5 headers survive into the PDF sheet
  * coalesced <col min=3 max=4> ranges keep column D (v2.1 regression guard)
  * fresh create_tab data rows are auto-height (long values grow, don't clip)
  * the Invoice Amount header is right-aligned; totals columns fit $-totals
  * page setup: landscape, repeat header row, fitToWidth None->1 but 0 preserved,
    a manual scale disables force-fit-to-page
"""
import os, sys, shutil, subprocess, zipfile, re
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import openpyxl
from openpyxl.drawing.image import Image as XlImage
import app

EMU = 9525
fails = []
def check(cond, msg):
    print(("  ok   " if cond else "  FAIL ") + msg)
    if not cond: fails.append(msg)

def drawing_ext_px(xlsx_path):
    z = zipfile.ZipFile(xlsx_path)
    for n in z.namelist():
        if 'drawing' in n and n.endswith('.xml'):
            xml = z.read(n).decode()
            ext = re.search(r'<(?:xdr:)?ext\s+cx="(\d+)"\s+cy="(\d+)"', xml)
            if ext:
                return int(ext.group(1))/EMU, int(ext.group(2))/EMU
    return None

def header_row_top_px(ws):
    return sum((ws.row_dimensions[r].height or 15) * 96/72 for r in range(1, 5))

def build_source(path, long_facility=False, big_total=False):
    """A minimal raw source: Invoice List + Dist Lookup for one distributor."""
    wb = openpyxl.Workbook(); il = wb.active; il.title = 'Invoice List'
    il['A3'] = 'June 2026'
    il['A6'] = 'D-001'
    facility = ('A Very Long Facility Name Medical Center of the Greater '
                'Metropolitan Region and Surrounding Counties') if long_facility \
               else 'Eaton Rapids Medical Center'
    amt = 98765432.10 if big_total else 4250.0
    comm = 29629629.63 if big_total else 1275.0
    il.cell(row=7, column=1, value='Mega Distributors LLC')
    for col, v in [(2, '06/03/2026'), (3, 'INV-1'), (4, 'PO-123'), (5, 'Dr X'),
                   (6, facility), (7, 'Widget'), (8, 0.30), (9, amt), (10, comm)]:
        il.cell(row=7, column=col, value=v)
    il.cell(row=8, column=1, value='Total for D-001')
    il.cell(row=8, column=9, value=amt); il.cell(row=8, column=10, value=comm)
    dl = wb.create_sheet('Dist Lookup')
    dl['A3'] = 'D-001'; dl['B3'] = 'Mega Distributors LLC'; dl['C3'] = 'mega@example.com'
    wb.save(path)

def make_review_wb(path, **kw):
    """Run Step 1 to get a real review workbook (Summary + distributor tab)."""
    job = '/tmp/_pdftest_s1'; shutil.rmtree(job, ignore_errors=True); os.makedirs(job)
    src = os.path.join(job, 'src.xlsx'); build_source(src, **kw)
    r = app.process_excel(src, job)
    shutil.copy(r['xlsx_path'], path)

def run_generate(review_xlsx):
    """Run generate_pdfs with soffice intercepted; return {tab_base: temp_xlsx}."""
    job = '/tmp/_pdftest_job'; shutil.rmtree(job, ignore_errors=True); os.makedirs(job)
    shutil.copy(review_xlsx, os.path.join(job, 'Commission_Statements_June_2026.xlsx'))
    kept = {}
    class R: returncode = 0; stderr = ''; stdout = ''
    def fake_run(*a, **k):
        cmd = a[0]; outdir = cmd[cmd.index('--outdir')+1]; infile = cmd[-1]
        base = os.path.splitext(os.path.basename(infile))[0]
        dst = f'/tmp/_pdftest_temp/{base}.xlsx'; os.makedirs('/tmp/_pdftest_temp', exist_ok=True)
        shutil.copy(infile, dst); kept[base] = dst
        open(os.path.join(outdir, base + '.pdf'), 'wb').write(b'%PDF-1.4\n%%EOF')
        return R()
    shutil.rmtree('/tmp/_pdftest_temp', ignore_errors=True)
    orig = subprocess.run; subprocess.run = fake_run
    try: res = app.generate_pdfs(job)
    finally: subprocess.run = orig
    return kept, res

def first_dist_tab(wb):
    skip = {'Summary', 'Invoice List', 'Trauma', 'Dist Lookup'}
    return next(s for s in wb.sheetnames if s not in skip)

# ---------------------------------------------------------------------------
print("=== TEST 1: logo never overlaps headers, across source logo sizes ===")
review = '/tmp/_pdftest_review.xlsx'; make_review_wb(review)
for w, h, label in [(453, 235, 'excel-inflated'), (900, 500, 'huge'),
                    (271, 125, 'native'), (220, 115, 'normal')]:
    wb = openpyxl.load_workbook(review); ws = wb[first_dist_tab(wb)]
    ws._images = []
    img = XlImage(app.LOGO_PATH); img.width = w; img.height = h; img.anchor = 'B1'
    ws.add_image(img)
    variant = f'/tmp/_pdftest_v.xlsx'; wb.save(variant)
    kept, _ = run_generate(variant)
    temp = next(iter(kept.values())); ws2 = openpyxl.load_workbook(temp).active
    ext = drawing_ext_px(temp)
    check(ext is not None and abs(ext[0]-220) < 2 and abs(ext[1]-115) < 2,
          f"[{label}] logo ext == 220x115 (got {ext})")
    check(ext and ext[1] < header_row_top_px(ws2),
          f"[{label}] logo bottom {ext[1]:.0f}px clears header row5 top {header_row_top_px(ws2):.0f}px")
    check(all(ws2.cell(row=5, column=c).value for c in range(2, 11)),
          f"[{label}] all 9 headers B5:J5 present")

# ---------------------------------------------------------------------------
print("\n=== TEST 2: column D preserved when C/D share a coalesced range ===")
wb = openpyxl.load_workbook(review); ws = wb[first_dist_tab(wb)]
# force C and D to identical width so a save coalesces them into <col min=3 max=4>
ws.column_dimensions['C'].width = 15.0; ws.column_dimensions['D'].width = 15.0
merged = '/tmp/_pdftest_merged.xlsx'; wb.save(merged)
kept, _ = run_generate(merged)
ws2 = openpyxl.load_workbook(next(iter(kept.values()))).active
check(ws2.column_dimensions['D'].width and ws2.column_dimensions['D'].width >= 14,
      f"col D width preserved (got {ws2.column_dimensions['D'].width})")

# ---------------------------------------------------------------------------
print("\n=== TEST 3: fresh Step-1 tab is safe (auto-height, aligns, totals, logo) ===")
review2 = '/tmp/_pdftest_review2.xlsx'; make_review_wb(review2, long_facility=True, big_total=True)
wb = openpyxl.load_workbook(review2); tab = first_dist_tab(wb); ws = wb[tab]
check(all(ws.cell(row=5, column=c).value for c in range(2, 11)), f"all headers present in {tab!r}")
check(ws.cell(row=5, column=9).alignment.horizontal == 'right', "Invoice Amount header right-aligned")
clip = [r for r, d in ws.row_dimensions.items() if d.height and abs(d.height-16.05) < 0.01 and r >= 7]
check(not clip, f"data rows are auto-height, no 16.05 clip (bad={clip})")
ext = drawing_ext_px(review2)
check(ext and abs(ext[0]-220) < 2 and abs(ext[1]-115) < 2, f"logo ext 220x115 (got {ext})")
need = len('$98,765,432.10')
check(ws.column_dimensions['I'].width >= need-1 and ws.column_dimensions['J'].width >= need-1,
      f"totals cols fit 8-figure $ total (I={ws.column_dimensions['I'].width:.1f} J={ws.column_dimensions['J'].width:.1f})")

# ---------------------------------------------------------------------------
print("\n=== TEST 4: page-setup hardening ===")
kept, _ = run_generate(review)
ws2 = openpyxl.load_workbook(next(iter(kept.values()))).active
check(ws2.page_setup.orientation == 'landscape', "orientation landscape")
check(str(ws2.print_title_rows) in ('5:5', '$5:$5'), "header row repeats on each page")
# fitToWidth=0 preserved; manual scale disables force fit-to-page
wb = openpyxl.load_workbook(review); ws = wb[first_dist_tab(wb)]
ws.page_setup.fitToWidth = 0; ws.page_setup.scale = 85
ws.sheet_properties.pageSetUpPr.fitToPage = False
ps = '/tmp/_pdftest_ps.xlsx'; wb.save(ps)
kept, _ = run_generate(ps); ws2 = openpyxl.load_workbook(next(iter(kept.values()))).active
check(ws2.page_setup.fitToWidth == 0, f"deliberate fitToWidth=0 preserved (got {ws2.page_setup.fitToWidth})")
check(ws2.page_setup.scale == 85, f"manual scale preserved (got {ws2.page_setup.scale})")
check(ws2.sheet_properties.pageSetUpPr.fitToPage in (False, None),
      "fit-to-page not force-enabled when a manual scale is set")

print("\n" + "=" * 52)
print("RESULT:", "ALL PASSED" if not fails else f"{len(fails)} FAILURE(S)")
for f in fails: print("  -", f)
sys.exit(1 if fails else 0)
