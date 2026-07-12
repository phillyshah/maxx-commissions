#!/usr/bin/env python3
"""Maxx Health Commission Statement Generator - Flask App"""

import os
import re
import zipfile
import subprocess
import uuid
import shutil
from datetime import datetime
from copy import copy

from flask import Flask, render_template, request, send_file, jsonify, url_for
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XlImage
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
app.config['OUTPUT_FOLDER'] = os.path.join(os.path.dirname(__file__), 'outputs')

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

# Version
APP_VERSION = "2.1"

# Release notes (newest first)
RELEASE_NOTES = [
    {
        'version': '2.1',
        'title': 'PDF Column Width & Row Clipping Fix',
        'description': 'Fixed a bug where the P.O. Number column collapsed to a narrow default width on some distributor PDFs, which made P.O. values wrap onto a second line and get clipped by the row height, and threw off the spacing next to the Surgeon column. Column widths are now copied faithfully from the reviewed workbook (including columns Excel groups together), so PDFs match the Excel layout.'
    },
    {
        'version': '2.0',
        'title': 'Download & Reliability Fixes',
        'description': 'Fixed a critical bug where clicking "Download Excel" after Step 1 could return the original uploaded source file instead of the generated review workbook. The download now always serves the correct Commission_Statements file. Also fixed the Summary sheet year (was hardcoded to 2026; now uses the year detected from the Invoice List), added a clear error if no distributor groups are found in the Invoice List, and PDF generation now reports a clear error if LibreOffice fails instead of silently producing an empty zip.'
    },
    {
        'version': '1.9',
        'title': 'Column Header Alignment Fix',
        'description': 'Column headers now center vertically in their row instead of being pinned to the top. Header row height also tightened for a cleaner, more professional look.'
    },
    {
        'version': '1.8',
        'title': 'Spacing Above Column Headers',
        'description': 'Added extra spacing between the rep contact info and the column header row so the table no longer looks scrunched.'
    },
    {
        'version': '1.7',
        'title': 'Header Text Wrapping Fix',
        'description': 'Column headers now always display from the top of the cell so wrapped text is never clipped. "Invoice Number" shortened to "Invoice Num". Invoice Date column width adjusted for clean display.'
    },
    {
        'version': '1.6',
        'title': 'P.O. Number Column Fix',
        'description': 'P.O. Number column now expands wide enough to show long values like "MAX040126Distal Radius855" without cutting off text. Name of Facility column also has a wider maximum.'
    },
    {
        'version': '1.5',
        'title': 'Auto-Fit Column Widths',
        'description': 'Column widths now automatically adjust to fit the actual content in each distributor tab. Long facility names and other text no longer get cut off.'
    },
    {
        'version': '1.4',
        'title': 'US Date Format in PDF Filenames',
        'description': 'PDF filenames now use US date format (MM-DD-YYYY) instead of DD Month YYYY. Example: S-Squared Surgical_Commission Statement Paid on 05-31-2026.pdf'
    },
    {
        'version': '1.3',
        'title': "What's New Feed",
        'description': 'Added a What\'s New link in the footer showing release history.'
    },
    {
        'version': '1.2',
        'title': 'PDF Generation Fixes',
        'description': 'Column headers now repeat on every page of multi-page statements. PDFs print in landscape orientation. User edits to column widths are preserved when generating PDFs. Invoice Date column now wraps correctly in print.'
    },
    {
        'version': '1.1',
        'title': 'Light Theme & Statement Formatting',
        'description': 'Switched to a light, high-contrast theme for easier reading. Logo resized. Distributorship name now appears above contact info. Invoice Date and Invoice Number left-justified. Column F renamed to Name of Facility. Column widths adjusted for better text fit. PDF files renamed with distributor name and payment date.'
    },
    {
        'version': '1.0',
        'title': 'Initial Release',
        'description': 'Two-step commission statement workflow: upload a source worksheet to generate a review workbook with Summary and distributor tabs, then upload the verified workbook to create a PDF zip bundle.'
    },
]

LOGO_PATH = os.path.join(os.path.dirname(__file__), 'static', 'maxx_logo.png')

# ─── Styles ───────────────────────────────────────────────────────────────────
FONT_DATE      = Font(name='Arial', size=10)
FONT_CONTACT   = Font(name='Arial', size=11)
FONT_DIST_LBL  = Font(name='Arial', size=11)
FONT_COMM_LBL  = Font(name='Arial', size=10)
FONT_HDR       = Font(name='Arial', size=9, bold=True)
FONT_CODE      = Font(name='Arial', size=9, bold=True)
FONT_DATA      = Font(name='Arial', size=9)
FONT_TOT_LBL   = Font(name='Arial', size=9, bold=True)
FONT_TOT_NUM   = Font(name='Arial', size=9, bold=True)
FONT_FOOTER    = Font(name='Arial', size=12, bold=True, italic=True)

FONT_SUM_TITLE = Font(name='Arial', size=12, bold=True)
FONT_SUM_HDR   = Font(name='Calibri', size=11)
FONT_SUM_HDR_A = Font(name='Calibri', size=12, bold=True)
FONT_SUM_DATA  = Font(name='Arial', size=10)
FONT_SUM_BOLD  = Font(name='Arial', size=10, bold=True)

W = True
ALIGN_HDR_C  = Alignment(horizontal='center', vertical='center', wrap_text=W)
ALIGN_HDR_L  = Alignment(horizontal='left',   vertical='center', wrap_text=W)
ALIGN_HDR_R  = Alignment(horizontal='right',  vertical='center', wrap_text=W)
ALIGN_DATA_L = Alignment(horizontal='left', wrap_text=W)
ALIGN_DATA_C = Alignment(horizontal='center', wrap_text=W)
ALIGN_DATA_R = Alignment(horizontal='right', wrap_text=W)
ALIGN_CODE_L = Alignment(horizontal='left', wrap_text=W)
ALIGN_RIGHT  = Alignment(horizontal='right')
ALIGN_CENTER = Alignment(horizontal='center')
ALIGN_CC     = Alignment(horizontal='center', vertical='center', wrap_text=W)
ALIGN_RC     = Alignment(horizontal='right', vertical='center')

BORDER_TOT_I  = Border(top=Side(style='thin'))
BORDER_TOT_J  = Border(top=Side(style='medium'), bottom=Side(style='medium'),
                       left=Side(style='medium'), right=Side(style='medium'))
BORDER_THIN_B = Border(bottom=Side(style='thin'))
BORDER_THIN_TB= Border(top=Side(style='thin'), bottom=Side(style='thin'))
BORDER_THIN_T = Border(top=Side(style='thin'))

COL_WIDTHS = {'A': 18.0, 'B': 9.5, 'C': 12.11, 'D': 14.22, 'E': 14.0,
              'F': 30.0, 'G': 24.0, 'H': 10.0, 'I': 15.89, 'J': 16.78}
# Min/max widths for auto-fit (columns B-J)
COL_MIN = {2: 11,  3: 12, 4: 14, 5: 10, 6: 18, 7: 12, 8: 7, 9: 13, 10: 12}
COL_MAX = {2: 14,  3: 18, 4: 35, 5: 20, 6: 50, 7: 35, 8: 10, 9: 16, 10: 16}
ROW_HEIGHTS = {1: 42.6, 2: 41.4, 3: 34.2, 4: 57.0, 5: 30.0}
DATA_ROW_H = 16.05
TOTAL_ROW_H = 18.6
SUM_COL_WIDTHS = {'A': 26.0, 'B': 17.89, 'C': 19.44, 'D': 13.11,
                  'E': 13.0, 'F': 15.89, 'G': 16.66}


# ─── Processing Functions ─────────────────────────────────────────────────────

def detect_month_year(ws):
    """Auto-detect month/year from the Invoice List header (row 3)."""
    for r in range(1, 6):
        for c in range(1, 10):
            val = ws.cell(row=r, column=c).value
            if val and isinstance(val, str):
                # Match patterns like "February 2026", "January 2026"
                m = re.match(r'(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})', val)
                if m:
                    month_name = m.group(1)
                    year = int(m.group(2))
                    month_num = ['January','February','March','April','May','June',
                                 'July','August','September','October','November','December'].index(month_name) + 1
                    return month_name, year, month_num
    return None, None, None


def compute_pay_date(year, month_num):
    """Payment date is last day of the month after the sales month."""
    import calendar
    pay_month = month_num + 1 if month_num < 12 else 1
    pay_year = year if month_num < 12 else year + 1
    last_day = calendar.monthrange(pay_year, pay_month)[1]
    return datetime(pay_year, pay_month, last_day)


def load_lookup(wb):
    ws = wb['Dist Lookup']
    lookup = {}
    for row in range(3, ws.max_row + 1):
        code = ws.cell(row=row, column=1).value
        name = ws.cell(row=row, column=2).value
        contact = ws.cell(row=row, column=3).value
        if code:
            lookup[str(code).strip()] = {
                'name': str(name).strip() if name else '',
                'contact': str(contact).strip() if contact else ''
            }
    return lookup


def parse_groups(ws):
    groups = []
    current_code = current_name = None
    current_data = []
    for row_num in range(6, ws.max_row + 1):
        a = ws.cell(row=row_num, column=1).value
        b = ws.cell(row=row_num, column=2).value
        if a and str(a).strip().startswith('Total for'):
            groups.append({'code': current_code, 'name': current_name, 'data': current_data,
                           'total_amount': ws.cell(row=row_num, column=9).value or 0,
                           'total_commission': ws.cell(row=row_num, column=10).value or 0})
            current_code = current_name = None
            current_data = []
        elif a and b is None and not str(a).strip().startswith('Total'):
            current_code = str(a).strip()
        elif b is not None:
            row_data = {col: ws.cell(row=row_num, column=col).value for col in range(1, 12)}
            if a and current_name is None:
                current_name = str(a).strip()
            current_data.append(row_data)
    return groups


def autofit_columns(ws, header_row, last_data_row):
    """Auto-fit columns B-J based on actual cell content."""
    col_letters = {2:'B', 3:'C', 4:'D', 5:'E', 6:'F', 7:'G', 8:'H', 9:'I', 10:'J'}
    for col_num in range(2, 11):
        max_len = 0
        for r in range(header_row, last_data_row + 1):
            val = ws.cell(row=r, column=col_num).value
            if val is None:
                continue
            if isinstance(val, datetime):
                text = val.strftime('%m/%d/%Y')
            elif isinstance(val, (int, float)):
                if col_num == 8:  # Rate
                    text = f'{val:.0%}' if val < 1 else f'{val}%'
                elif col_num >= 9:  # Amount/Commission
                    text = f'{val:,.2f}'
                else:
                    text = str(val)
            else:
                text = str(val)
            max_len = max(max_len, len(text))
        # Apply with multiplier for Arial 9pt, clamped to min/max
        width = max_len * 1.15 + 2
        mn = COL_MIN.get(col_num, 8)
        mx = COL_MAX.get(col_num, 40)
        ws.column_dimensions[col_letters[col_num]].width = max(mn, min(mx, width))


def create_tab(wb, tab_name, code, dist_name, contact, data_rows,
               total_amount, total_commission, pay_date, commission_label, logo_path):
    ws = wb.create_sheet(title=tab_name[:31])
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    for rn, h in ROW_HEIGHTS.items():
        ws.row_dimensions[rn].height = h

    # Logo
    if os.path.exists(logo_path):
        logo = XlImage(logo_path)
        logo.width = 220
        logo.height = 115
        logo.anchor = 'B1'
        ws.add_image(logo)

    # J1: date right-justified
    c = ws.cell(row=1, column=10, value=pay_date)
    c.font = FONT_DATE; c.number_format = 'm/d/yy;@'; c.alignment = ALIGN_RIGHT

    # B3: distributorship name first, J3: commission label right-justified
    ws.cell(row=3, column=2, value=f'Distributor:  {dist_name}').font = FONT_DIST_LBL
    c = ws.cell(row=3, column=10, value=commission_label)
    c.font = FONT_COMM_LBL; c.alignment = ALIGN_RIGHT

    # B4: contact name/email
    if contact:
        ws.cell(row=4, column=2, value=contact).font = FONT_CONTACT

    # Row 5: headers
    for col, text, align in [
        (2, 'Invoice Date', ALIGN_HDR_L), (3, 'Invoice Num', ALIGN_HDR_L),
        (4, 'P.O. Number', ALIGN_HDR_L), (5, 'Surgeon', ALIGN_HDR_L),
        (6, 'Name of Facility', ALIGN_HDR_L), (7, 'Memo/ Description', ALIGN_HDR_L),
        (8, 'Rate', ALIGN_HDR_C), (9, 'Invoice Amount', ALIGN_HDR_C),
        (10, 'Commission', ALIGN_HDR_R)]:
        c = ws.cell(row=5, column=col, value=text)
        c.font = FONT_HDR; c.alignment = align
    ws.cell(row=5, column=8).number_format = '0%'

    # Row 6: code
    ws.row_dimensions[6].height = DATA_ROW_H
    ws.cell(row=6, column=1, value=code).font = FONT_CODE
    ws.cell(row=6, column=1).alignment = ALIGN_CODE_L

    # Data rows
    row = 7
    for i, d in enumerate(data_rows):
        ws.row_dimensions[row].height = DATA_ROW_H
        if i == 0 and d.get(1):
            ws.cell(row=row, column=1, value=d[1]).font = FONT_CODE
        for col in range(2, 11):
            val = d.get(col)
            if val is not None:
                c = ws.cell(row=row, column=col, value=val)
                c.font = FONT_DATA
                if col <= 7: c.alignment = ALIGN_DATA_L
                elif col == 8: c.alignment = ALIGN_DATA_C; c.number_format = '0%'
                else: c.alignment = ALIGN_DATA_R; c.number_format = '#,##0.00\\ _€'
        row += 1

    # Total row
    ws.row_dimensions[row].height = TOTAL_ROW_H
    ws.cell(row=row, column=1, value=f'Total for {code}').font = FONT_TOT_LBL
    ws.cell(row=row, column=1).alignment = ALIGN_CODE_L
    for col, val, bdr in [(9, total_amount, BORDER_TOT_I), (10, total_commission, BORDER_TOT_J)]:
        c = ws.cell(row=row, column=col, value=val)
        c.font = FONT_TOT_NUM; c.alignment = ALIGN_DATA_R
        c.number_format = '"$"* #,##0.00\\ _€'; c.border = bdr

    # Footer
    footer_row = row + 2
    ws.merge_cells(start_row=footer_row, start_column=2, end_row=footer_row, end_column=10)
    c = ws.cell(row=footer_row, column=2, value='Thank you for your continued support.')
    c.font = FONT_FOOTER; c.alignment = ALIGN_CENTER

    # Auto-fit columns B-J to actual content
    autofit_columns(ws, 5, row - 1)  # row-1 = last data row (before total)

    # Page setup
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 1  # Letter
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f'B1:J{footer_row}'
    ws.print_title_rows = '5:5'  # Repeat header row on every page
    return ws


def create_summary(wb, groups, lookup, commission_label, year=None):
    ws = wb.create_sheet(title='Summary', index=0)
    for col, w in SUM_COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:G1'); ws.row_dimensions[1].height = 15.75
    c = ws.cell(row=1, column=1, value=year)
    c.font = FONT_SUM_TITLE; c.alignment = ALIGN_CENTER

    ws.merge_cells('A2:G2'); ws.row_dimensions[2].height = 15.75
    c = ws.cell(row=2, column=1, value=commission_label)
    c.font = FONT_SUM_TITLE; c.alignment = ALIGN_CENTER
    for col in range(1, 8):
        ws.cell(row=2, column=col).border = BORDER_THIN_B

    ws.row_dimensions[3].height = 35.4
    for col, text, font, align in [
        (1, 'Distributor', FONT_SUM_HDR_A, None),
        (2, 'Commission Earned', FONT_SUM_HDR, ALIGN_CC),
        (3, 'Chargeback to Maxx Orthopedics', FONT_SUM_HDR, ALIGN_CC),
        (4, 'Expense report payments', FONT_SUM_HDR, ALIGN_CC),
        (5, 'Other payments', FONT_SUM_HDR, ALIGN_CC),
        (6, 'Freight charges\ndeduction', FONT_SUM_HDR, ALIGN_CC),
        (7, 'Total Commission\nPaid', FONT_SUM_HDR, ALIGN_CC)]:
        c = ws.cell(row=3, column=col, value=text)
        c.font = font
        if align: c.alignment = align
        c.border = BORDER_THIN_TB

    sorted_groups = sorted(groups, key=lambda g: (lookup.get(g['code'], {}).get('name', g['name'] or '')).lower())
    row = 4
    total_all = 0
    for g in sorted_groups:
        ws.row_dimensions[row].height = 13.05
        info = lookup.get(g['code'], {})
        dist_name = info.get('name', g['name'] or '')
        ws.cell(row=row, column=1, value=dist_name).font = FONT_SUM_DATA
        c = ws.cell(row=row, column=2, value=g['total_commission'])
        c.font = FONT_SUM_DATA; c.number_format = '"$"#,##0.00'
        for cc in range(3, 7):
            ws.cell(row=row, column=cc).font = FONT_SUM_DATA
            ws.cell(row=row, column=cc).number_format = '"$"#,##0.00'
        c = ws.cell(row=row, column=7, value=g['total_commission'])
        c.font = FONT_SUM_DATA; c.number_format = '"$"#,##0.00'; c.alignment = ALIGN_RC
        total_all += g['total_commission']
        row += 1

    tr = row
    ws.row_dimensions[tr].height = 13.05
    ws.cell(row=tr, column=1, value='Total Distributor Commission:').font = FONT_SUM_DATA
    for col in range(1, 8):
        ws.cell(row=tr, column=col).border = BORDER_THIN_T
    for col in [2, 7]:
        c = ws.cell(row=tr, column=col, value=total_all)
        c.font = FONT_SUM_DATA; c.number_format = '"$"#,##0.00'

    r = tr + 3
    ws.cell(row=r, column=1, value='Total Commission').font = FONT_SUM_BOLD
    c = ws.cell(row=r, column=2, value=total_all)
    c.font = FONT_SUM_DATA; c.number_format = '"$"#,##0.00_);[Red]\\("$"#,##0.00\\)'
    ws.cell(row=r, column=6, value='Total Payments').font = FONT_SUM_DATA
    c = ws.cell(row=r, column=7, value=total_all)
    c.font = FONT_SUM_DATA; c.number_format = '"$"#,##0.00_);[Red]\\("$"#,##0.00\\)'
    r += 2
    ws.cell(row=r, column=6, value='Total ACH').font = FONT_SUM_DATA
    c = ws.cell(row=r, column=7, value=total_all)
    c.font = FONT_SUM_DATA; c.number_format = '"$"#,##0.00'
    r += 1
    ws.cell(row=r, column=6, value='Total Checks').font = FONT_SUM_DATA
    r += 1
    ws.cell(row=r, column=6, value='Total Payment').font = FONT_SUM_BOLD
    c = ws.cell(row=r, column=7, value=total_all)
    c.font = FONT_SUM_BOLD; c.number_format = '"$"#,##0.00'
    return ws


def process_excel(input_path, job_dir):
    """Step 1: Build the output Excel workbook. Returns xlsx path and metadata."""
    wb_src = openpyxl.load_workbook(input_path, data_only=True)

    # Validate required sheets
    required = {'Invoice List', 'Dist Lookup'}
    missing = required - set(wb_src.sheetnames)
    if missing:
        wb_src.close()
        raise ValueError(f"Missing required sheets: {', '.join(missing)}")

    # Auto-detect month/year
    month_name, year, month_num = detect_month_year(wb_src['Invoice List'])
    if not month_name:
        wb_src.close()
        raise ValueError("Could not detect month/year from Invoice List. Expected a row like 'February 2026'.")

    pay_date = compute_pay_date(year, month_num)
    commission_label = f'Commission on {month_name} {year} Sales'

    lookup = load_lookup(wb_src)
    groups = parse_groups(wb_src['Invoice List'])

    if not groups:
        wb_src.close()
        raise ValueError(
            "No distributor groups found in Invoice List. "
            "Expected data rows followed by 'Total for ...' summary rows starting at row 6."
        )

    # Build output workbook
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    create_summary(wb_out, groups, lookup, commission_label, year=year)

    # Copy source sheets
    for src_name in wb_src.sheetnames:
        src_ws = wb_src[src_name]
        new_ws = wb_out.create_sheet(title=src_name)
        for merge in src_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merge))
        for col_letter, dim in src_ws.column_dimensions.items():
            width = dim.width if dim.width else 8.43
            # Expand across the full min..max range so coalesced column ranges
            # (e.g. <col min=3 max=4>) don't drop columns. See generate_pdfs().
            start = dim.min or 1
            end = dim.max or start
            for ci in range(start, end + 1):
                new_ws.column_dimensions[get_column_letter(ci)].width = width
        for row_num, dim in src_ws.row_dimensions.items():
            new_ws.row_dimensions[row_num].height = dim.height if dim.height else 15
        for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, max_col=src_ws.max_column):
            for cell in row:
                if isinstance(cell, MergedCell): continue
                new_cell = new_ws.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = copy(cell.font); new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill); new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection); new_cell.alignment = copy(cell.alignment)

    wb_src.close()

    # Create distributor tabs sorted by name
    sorted_groups = sorted(groups, key=lambda x: (lookup.get(x['code'], {}).get('name', x['name'] or '')).lower())
    for g in sorted_groups:
        code = g['code']
        info = lookup.get(code, {})
        dist_name = info.get('name', g['name'] or '')
        contact = info.get('contact', '')
        tab_name = re.sub(r'[\\/*?\[\]:]', '', g['name'] or code)[:31]
        existing = [s.title for s in wb_out.worksheets]
        if tab_name in existing:
            tab_name = f"{tab_name[:27]} {code}"[:31]
        create_tab(wb_out, tab_name, code, dist_name, contact,
                   g['data'], g['total_amount'], g['total_commission'],
                   pay_date, commission_label, LOGO_PATH)

    # Save workbook
    xlsx_name = f'Commission_Statements_{month_name}_{year}.xlsx'
    xlsx_path = os.path.join(job_dir, xlsx_name)
    wb_out.save(xlsx_path)

    return {
        'xlsx_path': xlsx_path,
        'xlsx_name': xlsx_name,
        'month': month_name,
        'year': year,
        'num_distributors': len(groups),
    }


def validate_pdf_source_workbook(input_path):
    """Validate that the uploaded workbook is a reviewed statements workbook."""
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    sheetnames = set(wb.sheetnames)
    wb.close()

    if 'Summary' not in sheetnames:
        raise ValueError("The verified workbook must include a Summary sheet.")

    non_statement_sheets = {'Invoice List', 'Dist Lookup', 'Trauma', 'Summary'}
    distributor_tabs = [name for name in sheetnames if name not in non_statement_sheets]
    if not distributor_tabs:
        raise ValueError("The verified workbook must include at least one distributor tab.")


def generate_pdfs(job_dir):
    """Step 2: Convert the Excel workbook in job_dir to PDFs and zip them."""
    # Find the xlsx
    xlsx_path = None
    xlsx_name = None
    for fname in os.listdir(job_dir):
        if fname.endswith('.xlsx'):
            xlsx_path = os.path.join(job_dir, fname)
            xlsx_name = fname
            break
    if not xlsx_path:
        raise ValueError("No Excel workbook found for this job.")

    # Derive month/year from filename for zip name
    zip_base = xlsx_name.replace('.xlsx', '_PDFs')
    zip_name = f'{zip_base}.zip'
    zip_path = os.path.join(job_dir, zip_name)

    skip_sheets = {'Invoice List', 'Trauma', 'Dist Lookup', 'Summary'}
    temp_dir = os.path.join(job_dir, 'temp_sheets')
    pdf_dir = os.path.join(job_dir, 'pdfs')
    os.makedirs(temp_dir, exist_ok=True)
    os.makedirs(pdf_dir, exist_ok=True)

    wb_pdf = openpyxl.load_workbook(xlsx_path, data_only=True)
    for name in wb_pdf.sheetnames:
        if name in skip_sheets:
            continue
        src_ws = wb_pdf[name]
        # Build PDF filename: "DistName_Commission Statement Paid on April 30, 2026"
        pay_date_val = src_ws.cell(row=1, column=10).value
        if isinstance(pay_date_val, datetime):
            date_str = pay_date_val.strftime('%m-%d-%Y')
        else:
            date_str = 'Unknown Date'
        raw_file = f"{name}_Commission Statement Paid on {date_str}"
        safe_name = re.sub(r'[<>:"/\\|?*]', '_', raw_file).strip()
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = name[:31]
        for merge in src_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merge))
        for col_letter, dim in src_ws.column_dimensions.items():
            width = dim.width if dim.width else 8.43
            # A ColumnDimension can span a RANGE (Excel coalesces adjacent
            # columns of equal width into one <col min=3 max=4>). openpyxl keys
            # that range only under its min column, so copying per-key drops the
            # other columns and they collapse to the default width in the PDF —
            # this is what narrowed the P.O. Number column and caused the
            # wrap/clip and bad Surgeon spacing. Apply the width to EVERY column
            # in the range.
            start = dim.min or 1
            end = dim.max or start
            for ci in range(start, end + 1):
                L = get_column_letter(ci)
                new_ws.column_dimensions[L].width = width
                new_ws.column_dimensions[L].hidden = dim.hidden
        for row_num, dim in src_ws.row_dimensions.items():
            new_ws.row_dimensions[row_num].height = dim.height if dim.height else 15
            new_ws.row_dimensions[row_num].hidden = dim.hidden
        for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, max_col=src_ws.max_column):
            for cell in row:
                if isinstance(cell, MergedCell): continue
                new_cell = new_ws.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = copy(cell.font); new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill); new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection); new_cell.alignment = copy(cell.alignment)
        for img in src_ws._images:
            if os.path.exists(LOGO_PATH):
                new_img = XlImage(LOGO_PATH)
                new_img.width = img.width; new_img.height = img.height
                new_img.anchor = 'B1'
                new_ws.add_image(new_img)
        # Copy print settings from source (preserves user edits)
        if src_ws.print_area:
            new_ws.print_area = src_ws.print_area
        if src_ws.print_title_rows:
            new_ws.print_title_rows = src_ws.print_title_rows
        if src_ws.print_title_cols:
            new_ws.print_title_cols = src_ws.print_title_cols
        # Page setup — copy from source, fall back to landscape Letter
        new_ws.page_setup.orientation = src_ws.page_setup.orientation or 'landscape'
        new_ws.page_setup.paperSize = src_ws.page_setup.paperSize or 1  # 1 = Letter
        new_ws.page_setup.fitToWidth = src_ws.page_setup.fitToWidth if src_ws.page_setup.fitToWidth else 1
        new_ws.page_setup.fitToHeight = src_ws.page_setup.fitToHeight if src_ws.page_setup.fitToHeight is not None else 0
        if src_ws.page_setup.scale:
            new_ws.page_setup.scale = src_ws.page_setup.scale
        new_ws.sheet_properties.pageSetUpPr.fitToPage = True
        new_wb.save(os.path.join(temp_dir, f"{safe_name}.xlsx"))
        new_wb.close()

    wb_pdf.close()

    # Convert to PDF.
    # NOTE: soffice returns exit code 0 even when it fails to load/convert a file
    # (e.g. "Error: source file could not be loaded"), so we can't trust the
    # return code alone — we must verify the expected .pdf was actually written
    # and surface soffice's output if it wasn't. Otherwise a failed conversion
    # silently produces an empty zip and the UI reports a false success.
    lo_home = os.path.join(job_dir, 'lo_home')
    os.makedirs(lo_home, exist_ok=True)
    failures = []
    expected = 0
    for fname in sorted(os.listdir(temp_dir)):
        if not fname.endswith('.xlsx'): continue
        expected += 1
        expected_pdf = os.path.join(pdf_dir, fname[:-5] + '.pdf')
        try:
            proc = subprocess.run(
                ['soffice', '--headless', '--norestore', '--calc',
                 '--convert-to', 'pdf', '--outdir', pdf_dir, os.path.join(temp_dir, fname)],
                capture_output=True, text=True, timeout=120,
                env={**os.environ, 'HOME': lo_home})
            if not os.path.exists(expected_pdf):
                detail = (proc.stderr or proc.stdout or '').strip().splitlines()
                detail = detail[-1] if detail else f'exit code {proc.returncode}'
                failures.append(f'{fname[:-5]}: {detail}')
        except subprocess.TimeoutExpired:
            failures.append(f'{fname[:-5]}: conversion timed out')

    num_pdfs = len([f for f in os.listdir(pdf_dir) if f.endswith('.pdf')])

    # If nothing converted, fail loudly rather than returning an empty zip.
    if expected and num_pdfs == 0:
        shutil.rmtree(temp_dir, ignore_errors=True)
        shutil.rmtree(lo_home, ignore_errors=True)
        msg = "PDF conversion failed — LibreOffice could not generate any PDFs."
        if failures:
            msg += " First error: " + failures[0]
        msg += " Check that LibreOffice (soffice) is installed and working on the server."
        raise ValueError(msg)

    # Zip PDFs
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for pdf in sorted(os.listdir(pdf_dir)):
            if pdf.endswith('.pdf'):
                zf.write(os.path.join(pdf_dir, pdf), pdf)

    # Clean up temp
    shutil.rmtree(temp_dir, ignore_errors=True)
    shutil.rmtree(lo_home, ignore_errors=True)

    return {
        'zip_path': zip_path,
        'zip_name': zip_name,
        'num_pdfs': num_pdfs,
        'partial_failures': failures,
    }


# ─── Routes ───────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html', version=APP_VERSION, release_notes=RELEASE_NOTES)


@app.route('/upload', methods=['POST'])
@app.route('/upload-template', methods=['POST'])
def upload_template():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Please upload an .xlsx file'}), 400

    # Create job directory
    job_id = str(uuid.uuid4())[:8]
    job_dir = os.path.join(app.config['OUTPUT_FOLDER'], job_id)
    os.makedirs(job_dir, exist_ok=True)

    # Save uploaded file
    input_path = os.path.join(job_dir, secure_filename(file.filename))
    file.save(input_path)

    try:
        result = process_excel(input_path, job_dir)
        return jsonify({
            'success': True,
            'job_id': job_id,
            'month': result['month'],
            'year': result['year'],
            'num_distributors': result['num_distributors'],
            'xlsx_name': result['xlsx_name'],
        })
    except Exception as e:
        shutil.rmtree(job_dir, ignore_errors=True)
        return jsonify({'error': str(e)}), 500


@app.route('/generate-pdfs-upload', methods=['POST'])
def generate_pdfs_upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Please upload an .xlsx file'}), 400

    job_id = str(uuid.uuid4())[:8]
    job_dir = os.path.join(app.config['OUTPUT_FOLDER'], job_id)
    os.makedirs(job_dir, exist_ok=True)

    input_name = secure_filename(file.filename)
    input_path = os.path.join(job_dir, input_name)
    file.save(input_path)

    try:
        validate_pdf_source_workbook(input_path)
        result = generate_pdfs(job_dir)
        return jsonify({
            'success': True,
            'job_id': job_id,
            'source_name': input_name,
            'zip_name': result['zip_name'],
            'num_pdfs': result['num_pdfs'],
        })
    except Exception as e:
        shutil.rmtree(job_dir, ignore_errors=True)
        return jsonify({'error': str(e)}), 500


@app.route('/generate-pdfs/<job_id>', methods=['POST'])
def generate_pdfs_route(job_id):
    job_dir = os.path.join(app.config['OUTPUT_FOLDER'], job_id)
    if not os.path.exists(job_dir):
        return jsonify({'error': 'Job not found'}), 404

    try:
        result = generate_pdfs(job_dir)
        return jsonify({
            'success': True,
            'zip_name': result['zip_name'],
            'num_pdfs': result['num_pdfs'],
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/download/<job_id>/<filetype>')
def download(job_id, filetype):
    job_dir = os.path.join(app.config['OUTPUT_FOLDER'], job_id)
    if not os.path.exists(job_dir):
        return 'File not found', 404

    for fname in sorted(os.listdir(job_dir)):
        # For xlsx, only serve the generated output — never the uploaded source file.
        # Generated files always start with "Commission_Statements_".
        if filetype == 'xlsx' and fname.startswith('Commission_Statements_') and fname.endswith('.xlsx'):
            return send_file(os.path.join(job_dir, fname), as_attachment=True, download_name=fname)
        if filetype == 'zip' and fname.endswith('.zip'):
            return send_file(os.path.join(job_dir, fname), as_attachment=True, download_name=fname)

    return 'File not found', 404


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
